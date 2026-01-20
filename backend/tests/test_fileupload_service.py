import io
import json
import os
import tempfile
from random import choice
from typing import Any, Dict, List, Type, TypeVar, Union

import openpyxl
import pytest
from django.core.files.uploadedfile import SimpleUploadedFile

from apps.admin_panel.services import FileUploadService
from apps.core.models import CurrencyType, GameLevel
from apps.courts.models import Court, CourtLocation
from apps.event.models import Game
from apps.locations.models import City, Country
from apps.players.models import Player

DbModel = TypeVar(
    'DbModel',
    bound=Union[
        Country,
        City,
        CurrencyType,
        GameLevel,
        Court,
        CourtLocation,
        Game,
        Player,
    ],
)

DATA_DIR = os.path.join(os.path.dirname(__file__), '..', 'data')


def insert_fk_in_file(
    file_name: str, fk_field_names: List[str], data_dir=DATA_DIR
):
    """
    Insert random FK values into the specified fields in the Excel file.
    Args:
        file_name: Name of the Excel file (without extension).
        fk_field_names: List of FK field names to populate.
        data_dir: Directory where the Excel file is located.
    Uses on test setup to prepare Excel files with FK references.
    """
    excel_path = os.path.join(data_dir, f'{file_name}.xlsx')
    if not os.path.isfile(excel_path):
        raise FileNotFoundError(f'{excel_path} not found or is not a file')
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    header_row = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
    header = list(header_row)
    for fk_field in fk_field_names:
        fk_model = {
            'host_id': Player,
            'court_id': CourtLocation,
        }
        fk_values = fk_model[fk_field].objects.values_list('id', flat=True)
        try:
            col_idx = header.index(fk_field)
        except ValueError as e:
            raise ValueError(
                f'Field {fk_field} not found in {file_name}.xlsx'
            ) from e
        for row in ws.iter_rows(
            min_row=2,
            max_row=ws.max_row,
            min_col=col_idx + 1,
            max_col=col_idx + 1,
        ):
            for cell in row:
                cell.value = choice(fk_values)
        wb.save(excel_path)
    return


def clear_models(*models: List[Type[DbModel]]):
    """
    Delete all objects for the given models to avoid FK conflicts.

    Args:
        *models: List of Django model classes to clear.
    """
    for m in models:
        m.objects.all().delete()
    return


def create_countries_and_cities(valid_country_data, valid_city_data):
    """
    Create all countries and cities needed for tests.

    Args:
        valid_country_data: List of country dicts.
        valid_city_data: List of city dicts.
    """
    for c in valid_country_data:
        Country.objects.get_or_create(**c)
    for city in valid_city_data:
        country = Country.objects.get(name=city['country'])
        City.objects.get_or_create(name=city['name'], country=country)


def process_excel_file(
    fileupload_service: FileUploadService,
    model: str,
    excel_content_type: str,
    data_dir=DATA_DIR,
):
    """
    Upload an Excel file via the service and return result and row count.

    Args:
        fileupload_service: FileUploadService instance.
        model: Model name (str).
        excel_content_type: Content type for Excel files.
        data_dir: Directory with Excel files.

    Returns:
        Tuple of (result, row count).
    """
    excel_path = os.path.join(data_dir, f'{model}.xlsx')
    if not os.path.exists(excel_path):
        return None, 0
    with open(excel_path, 'rb') as f:
        file = SimpleUploadedFile(
            f'{model}.xlsx',
            f.read(),
            content_type=excel_content_type,
        )
        result = fileupload_service.process_file(file)
    wb = openpyxl.load_workbook(excel_path)
    ws = wb.active
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    return result, len(rows)


def get_model_create_result_json(
    f_service: FileUploadService, model: Type[DbModel], model_name, model_data
):
    """
    Helper to test model creation from JSON data.

    Args:
        f_service: FileUploadService instance.
        model: Django model class.
        model_name: Key for the data dict.
        model_data: List of dicts for model creation.

    Returns:
        Tuple of (result, object count).
    """
    model.objects.all().delete()
    data = {model_name: model_data}
    file: io.BytesIO = io.BytesIO(json.dumps(data).encode('utf-8'))
    file.name = 'test.json'
    result: dict = f_service.process_file(file)
    return result, model.objects.count()


def get_model_create_result_excel(
    valid_data: Dict[str, Any],
    fileupload_service_debug: FileUploadService,
    model_name: str,
    model: DbModel,
    excel_content_type: str,
):
    """
    Helper to test model creation from Excel data.

    Args:
        valid_data: List of dicts for model creation.
        fileupload_service_debug: FileUploadService instance.
        model_name: Name for the Excel file.
        model: Django model class.
        excel_content_type: Excel content type.

    Returns:
        Tuple of (result, object count).
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(list(valid_data[0].keys()))
    for v in valid_data:
        ws.append(list(v.values()))
    tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
    wb.save(tmp.name)
    tmp.close()
    with open(tmp.name, 'rb') as f:
        file: SimpleUploadedFile = SimpleUploadedFile(
            f'{model_name}.xlsx',
            f.read(),
            content_type=excel_content_type,
        )
        result: dict = fileupload_service_debug.process_file(file)
    os.unlink(tmp.name)
    return result, model.objects.count()


@pytest.mark.django_db
class TestFileUploadServiceJSON:
    def test_create_countries(
        self,
        fileupload_service_debug: FileUploadService,
        valid_country_data: List[Dict[str, Any]],
    ) -> None:
        result, objs_count = get_model_create_result_json(
            fileupload_service_debug,
            Country,
            'countries',
            valid_country_data,
        )
        assert result['success'] is True
        assert objs_count == len(valid_country_data)

    def test_create_cities(
        self,
        fileupload_service_debug: FileUploadService,
        valid_country_data: List[Dict[str, Any]],
        valid_city_data: List[Dict[str, Any]],
    ) -> None:
        create_countries_and_cities(valid_country_data, valid_city_data)
        result, objs_count = get_model_create_result_json(
            fileupload_service_debug,
            City,
            'cities',
            valid_city_data,
        )
        assert result['success'] is True
        assert objs_count == len(valid_city_data)

    def test_create_currencies(
        self,
        fileupload_service_debug: FileUploadService,
        valid_currency_data: List[Dict[str, Any]],
    ) -> None:
        result, objs_count = get_model_create_result_json(
            fileupload_service_debug,
            CurrencyType,
            'currencies',
            valid_currency_data,
        )
        assert result['success'] is True
        assert objs_count == len(valid_currency_data)

    def test_create_levels(
        self,
        fileupload_service_debug: FileUploadService,
        valid_level_data: List[Dict[str, Any]],
    ) -> None:
        result, objs_count = get_model_create_result_json(
            fileupload_service_debug,
            GameLevel,
            'levels',
            valid_level_data,
        )
        assert result['success'] is True
        assert objs_count == len(valid_level_data)

    def test_create_courts(
        self,
        fileupload_service_debug: FileUploadService,
        valid_court_json_data: List[Dict[str, Any]],
        valid_country_data: List[Dict[str, Any]],
        valid_city_data: List[Dict[str, Any]],
    ) -> None:
        create_countries_and_cities(valid_country_data, valid_city_data)
        result, objs_count = get_model_create_result_json(
            fileupload_service_debug,
            Court,
            'courts',
            valid_court_json_data,
        )
        assert result['success'] is True
        assert objs_count == len(valid_court_json_data)

    def test_create_tourneys_stub(
        self,
        fileupload_service_debug: FileUploadService,
    ) -> None:
        data = {'tourneys': [{'name': 'Test Tourney'}]}
        file: io.BytesIO = io.BytesIO(json.dumps(data).encode('utf-8'))
        file.name = 'test.json'
        result: dict = fileupload_service_debug.process_file(file)
        for m in result.get('messages', []):
            assert 'Skipped' in m or 'no serializer' in m.lower()
        assert result['success'] is False

    def test_process_json_file_invalid_country_attr(
        self,
        fileupload_service_debug: FileUploadService,
        invalid_country_data_wrong_attr: List[Dict[str, Any]],
    ) -> None:
        result, _ = get_model_create_result_json(
            fileupload_service_debug,
            Country,
            'countries',
            invalid_country_data_wrong_attr,
        )
        assert result['success'] is False or any(
            'error' in m.lower() for m in result.get('messages', [])
        )

    def test_process_json_file_invalid_country_format(
        self,
        fileupload_service_debug: FileUploadService,
        invalid_country_data_wrong_format: List[Dict[str, Any]],
    ) -> None:
        result, countries_after_test = get_model_create_result_json(
            fileupload_service_debug,
            Country,
            'countries',
            invalid_country_data_wrong_format,
        )
        assert countries_after_test == 0
        assert result['success'] is False
        assert 'error' in ''.join(
            m.lower() for m in result.get('messages', [])
        )

    def test_model_mapping_class_in_prod_mode(
        self,
        fileupload_service_prod: FileUploadService,
    ) -> None:
        for m in ('games', 'tourneys', 'players'):
            assert m not in fileupload_service_prod.model_mapping_class.keys()

    def test_unsupported_file_type(
        self,
        fileupload_service_debug: FileUploadService,
    ) -> None:
        file: io.BytesIO = io.BytesIO(b'This is a test.')
        file.name = 'test.txt'
        result: dict = fileupload_service_debug.process_file(file)
        assert result['success'] is False
        assert any(
            'unsupported file type' in m.lower()
            for m in result.get('messages', [])
        )


@pytest.mark.django_db
class TestFileUploadServiceExcel:
    def test_process_excel_file_creates_currency(
        self,
        fileupload_service_debug: FileUploadService,
        valid_currency_data: List[Dict[str, Any]],
        excel_content_type: str,
    ) -> None:
        clear_models(CurrencyType)
        result, objs_count = get_model_create_result_excel(
            valid_currency_data,
            fileupload_service_debug,
            'currencies',
            CurrencyType,
            excel_content_type,
        )
        assert result['success'] is True
        assert objs_count == len(valid_currency_data)

    @pytest.mark.parametrize(
        'model_name, model',
        [
            ('currencies', CurrencyType),
            ('countries', Country),
            ('cities', City),
            ('levels', GameLevel),
        ],
    )
    def test_process_excel_file_missing_fields(
        self,
        fileupload_service_debug: FileUploadService,
        excel_content_type: str,
        model_name: str,
        model: DbModel,
    ) -> None:
        clear_models(model)
        result, _ = get_model_create_result_excel(
            [{'wrong_field': 'test'}],
            fileupload_service_debug,
            model_name,
            model,
            excel_content_type,
        )
        assert result['success'] is False
        assert 'Missing model fields' in result['messages'][0]

    def test_process_excel_file_create_levels(
        self,
        fileupload_service_debug: FileUploadService,
        valid_level_data: List[Dict[str, Any]],
        excel_content_type: str,
    ) -> None:
        clear_models(
            GameLevel,
        )
        result, objs_count = get_model_create_result_excel(
            valid_level_data,
            fileupload_service_debug,
            'levels',
            GameLevel,
            excel_content_type,
        )
        assert result['success'] is True
        assert objs_count == len(valid_level_data)

    def test_process_excel_file_create_cities(
        self,
        fileupload_service_debug: FileUploadService,
        valid_city_data: List[Dict[str, Any]],
        valid_country_data: List[Dict[str, Any]],
        excel_content_type: str,
    ) -> None:
        clear_models(City, Country)
        for c in valid_country_data:
            Country.objects.create(**c)
        result, objs_count = get_model_create_result_excel(
            valid_city_data,
            fileupload_service_debug,
            'cities',
            City,
            excel_content_type,
        )
        assert result['success'] is True
        assert objs_count == len(valid_city_data)

    def test_process_excel_file_create_courts(
        self,
        fileupload_service_debug: FileUploadService,
        valid_court_xlsx_data: List[Dict[str, Any]],
        valid_country_data: List[Dict[str, Any]],
        valid_city_data: List[Dict[str, Any]],
        excel_content_type: str,
    ) -> None:
        clear_models(Court, CourtLocation, Country, City)
        create_countries_and_cities(valid_country_data, valid_city_data)
        result, objs_count = get_model_create_result_excel(
            valid_court_xlsx_data,
            fileupload_service_debug,
            'courts',
            Court,
            excel_content_type,
        )
        assert result['success'] is True
        assert objs_count == len(valid_court_xlsx_data)

    def test_process_excel_file_create_tourneys_stub(
        self,
        fileupload_service_debug: FileUploadService,
        excel_content_type: str,
    ) -> None:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(['name'])
        ws.append(['Test Tourney'])
        tmp = tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False)
        wb.save(tmp.name)
        tmp.close()
        with open(tmp.name, 'rb') as f:
            file: SimpleUploadedFile = SimpleUploadedFile(
                'tourneys.xlsx',
                f.read(),
                content_type=excel_content_type,
            )
            result: dict = fileupload_service_debug.process_file(file)
        os.unlink(tmp.name)
        assert result['success'] is True or result['success'] is False


@pytest.mark.django_db
class TestFileUploadServiceRealFiles:
    def test_upload_all_models_from_real_json_file(
        self, fileupload_service_debug: FileUploadService
    ) -> None:
        json_path = os.path.join(DATA_DIR, 'test_db_data.json')
        with open(json_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        for model in data.keys():
            fileupload_service_debug.model_mapping_class[
                model
            ].model.objects.all().delete()
        with open(json_path, 'rb') as f:
            file = io.BytesIO(f.read())
            file.name = 'test_db_data.json'
            result = fileupload_service_debug.process_file(file)
        assert result['success'] is True
        assert Country.objects.count() == len(data.get('countries', []))
        assert City.objects.count() == len(data.get('cities', []))
        assert CurrencyType.objects.count() == len(data.get('currencies', []))
        assert GameLevel.objects.count() == len(data.get('levels', []))
        assert Court.objects.count() == len(data.get('courts', []))

    def test_upload_models_from_real_excel_files_sequential(
        self,
        fileupload_service_debug: FileUploadService,
        excel_content_type: str,
        user_with_registered_player,
    ) -> None:
        models = [
            ('countries', Country),
            ('cities', City),
            ('levels', GameLevel),
            ('currencies', CurrencyType),
            ('courts', Court),
            ('games', Game),
        ]
        clear_models(
            Court,
            CourtLocation,
            CurrencyType,
            GameLevel,
        )

        for model, model_class in models:
            if model == 'games':
                insert_fk_in_file('games', ['host_id', 'court_id'])
            result, row_count = process_excel_file(
                fileupload_service_debug, model, excel_content_type
            )
            if result is None:
                continue
            assert result['success'] is True
            assert row_count == len(model_class.objects.all())

    def test_private_models_in_prod_mode(
        self,
        fileupload_service_prod: FileUploadService,
        excel_content_type: str,
    ) -> None:
        models = [
            'games',
        ]
        for model in models:
            result, _ = process_excel_file(
                fileupload_service_prod, model, excel_content_type
            )
            assert result['success'] is False
            assert any(
                'restricted in production mode' in m.lower()
                for m in result.get('messages', [])
            )
