import json
import os

import openpyxl

from apps.admin_panel.constants import UploadServiceMessages
from apps.admin_panel.exceptions import (
    ExcelValidationError,
    MissingSerializerError,
    RestrictedFileError,
    UnknownMappingError,
)
from apps.admin_panel.model_mappings import BaseModelMapping, MappingRegistry
from apps.core.base import BaseParser


class JsonParser(BaseParser):
    """
    Parser for JSON formatted files.
    """

    def parse(self, file):
        """
        Reads and decodes a JSON file into a Python dictionary.

        Args:
            file: A file-like object containing JSON data.

        Returns:
            dict: Parsed JSON data.
        """
        file.seek(0)
        return json.loads(file.read().decode('utf-8'))


class ExcelParser(BaseParser):
    """
    Parser for Excel (.xlsx) files using a mapping registry.
    """

    def parse(self, file, mapping_registry: MappingRegistry):
        """
        Parses an Excel file based on its filename and mappings.

        Args:
            file: An Excel file-like object.

        Returns:
            dict: {filename: list_of_aggregated_data}.

        Raises:
            RestrictedFileError: If the file is marked as private.
            UnknownMappingError: If no mapping is found for filename.
            MissingSerializerError: If mapping lacks a serializer.
            ExcelValidationError: If headers do not match expected fields.
        """
        filename = os.path.splitext(file.name)[0].lower()
        mapping_class: BaseModelMapping = mapping_registry.get_mapping(
            filename
        )
        if not mapping_class:
            if mapping_registry.is_private(filename):
                raise RestrictedFileError(
                    f'{UploadServiceMessages.RESTRICTED_UPLOAD}{filename}'
                )
            raise UnknownMappingError(
                f'{UploadServiceMessages.UNKNOWN_MODEL_MAPPING}{filename}'
            )
        if mapping_class.serializer is None:
            raise MissingSerializerError(
                f'{UploadServiceMessages.NO_MODEL_SERIALIZER}{filename}'
            )
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [
            cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        excel_fields = set(headers)
        expected_fields = set(mapping_class.expected_fields)
        if excel_fields != expected_fields:
            if expected_fields.issuperset(excel_fields):
                missing = expected_fields - excel_fields
                msg = (
                    f'{UploadServiceMessages.EXCEL_MISSING_MODEL_FIELDS}'
                    f'{list(missing)}'
                )
            else:
                invalid = excel_fields - expected_fields
                msg = (
                    f'{UploadServiceMessages.EXCEL_INVALID_MODEL_FIELDS}'
                    f'{list(invalid)}'
                )
            raise ExcelValidationError(msg)
        model_data = [
            mapping_class.agregate_model_fields(
                dict(zip(headers, row, strict=False))
            )
            for row in ws.iter_rows(min_row=2, values_only=True)
        ]
        return {filename: model_data}
