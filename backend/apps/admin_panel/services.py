import json
import logging
import os
from typing import Any, Dict, List

import openpyxl
from django.conf import settings

from apps.admin_panel.constants import (
    MAX_FILE_SIZE,
    SUPPORTED_FILE_TYPES,
    UploadServiceMessages,
)
from apps.admin_panel.model_mappings import (
    BaseModelMapping,
    CityModelMapping,
    CountryModelMapping,
    CourtModelMapping,
    CurrencyTypeModelMapping,
    GameModelMapping,
    LevelsModelMapping,
    PlayerModelMapping,
    TourneyModelMapping,
)

logger = logging.getLogger(__name__)


class FileUploadService:
    """
    Service for processing uploaded files.
    Supports JSON file type for loading model data.
    Implements singleton pattern.
    attrs:
        _model_mapping_class (Dict[str, Any]): Mapping of model names to their
            handlers.
        _model_processing_order (tuple[str]): Order in which models should be
            processed.
        _private_models_mapping_class (Dict[str, Any]): Mapping for private
            models only accessible in debug mode.
        max_file_size (int): Maximum allowed file size for upload.
        _supported_file_types (tuple[str]): Supported file types for upload.
        _extended_model_access (bool): Enable extended model processing.
    """

    def __new__(cls):
        """Singleton pattern implementation."""
        if not hasattr(cls, 'instance'):
            cls.instance = super(FileUploadService, cls).__new__(cls)
        return cls.instance

    def __init__(self):
        """Initialize model mapping for file processing."""
        self._model_mapping_class = {
            'countries': CountryModelMapping(),
            'cities': CityModelMapping(),
            'courts': CourtModelMapping(),
            'currencies': CurrencyTypeModelMapping(),
            'levels': LevelsModelMapping(),
        }
        self._private_models_mapping_class = {
            'players': PlayerModelMapping(),
            'games': GameModelMapping(),
            'tourneys': TourneyModelMapping(),
        }
        self._model_processing_order: tuple[str] = (
            'levels',
            'countries',
            'cities',
            'currencies',
            'courts',
        )
        self._supported_file_types: tuple[str] = SUPPORTED_FILE_TYPES
        self.max_file_size: int = MAX_FILE_SIZE
        self._extended_model_access: bool = settings.DEBUG

    @property
    def model_mapping_class(self) -> Dict[str, Any]:
        if self._extended_model_access:
            return {
                **self._model_mapping_class,
                **self._private_models_mapping_class,
            }
        return self._model_mapping_class

    @property
    def model_processing_order(self) -> tuple[str]:
        if self._extended_model_access:
            return self._model_processing_order + tuple(
                self._private_models_mapping_class.keys()
            )
        return self._model_processing_order

    @property
    def private_models_mapping_class(self) -> Dict[str, Any]:
        return self._private_models_mapping_class

    @property
    def supported_file_types(self) -> tuple[str]:
        return self._supported_file_types

    def detect_file_type(self, file) -> str:
        """Automatically detect file type."""
        filename = file.name.lower()
        if filename.endswith('.json'):
            return 'json'
        if filename.endswith('.csv'):
            return 'csv'
        if filename.endswith('.xlsx') or filename.endswith('.xls'):
            return 'excel'
        return filename.split('.')[-1]

    def process_file(self, file) -> Dict[str, Any]:
        """Process file based on its type."""
        file_type = self.detect_file_type(file)
        if file_type == 'json':
            return self.proccess_json_load(file)
        if file_type == 'excel':
            return self._process_excel_file(file)
        return {
            'success': False,
            'messages': [
                UploadServiceMessages.FILE_TYPE_NOT_SUPPORTED + file_type
            ],
        }

    def download_file_by_path(self, file_path: str) -> bytes:
        """Download file from given path."""
        if not os.path.exists(file_path):
            raise FileNotFoundError(f'file not found at {file_path}')
        with open(file_path, 'r', encoding='utf-8') as f:
            return f.read()

    def proccess_json_load(
        self,
        file,
    ) -> Dict[str, Any]:
        """Process JSON file with model data."""
        try:
            logger.info('Processing JSON file upload')
            file.seek(0)
            content = file.read().decode('utf-8')
            data = json.loads(content)
            available_models = list(data.keys())
            if not available_models:
                logger.error(UploadServiceMessages.NO_DATA_IN_JSON)
                return {
                    'success': False,
                    'messages': [UploadServiceMessages.NO_DATA_IN_JSON],
                }
            messages = []
            for model_name in self.model_processing_order:
                if model_name in data and data[model_name]:
                    if self.model_mapping_class[model_name].serializer is None:
                        m = f'Skipped - {model_name} with no serializer'
                        if model_name in self.private_models_mapping_class:
                            m = (
                                f'Skipped - {model_name} private model'
                                'in production mode'
                            )
                        logger.warning(m)
                        messages.append(m)
                        continue
                    result = self._process_model_data(
                        model_name,
                        data[model_name],
                    )
                    messages.extend(result['messages'])
            return {
                'success': any('created' in msg.lower() for msg in messages),
                'messages': messages,
            }
        except Exception as e:
            message = f'Error processing JSON file: {str(e)}'
            logger.error(message)
            return {
                'success': False,
                'messages': [message],
            }

    def _process_model_data(
        self,
        model_name: str,
        model_data: List[Dict],
    ) -> Dict[str, Any]:
        """Process data for specific model using serializer validation."""
        mapping_class: BaseModelMapping = self.model_mapping_class[model_name]
        messages: list[dict[str]] = []
        for obj_data in model_data:
            serializer = mapping_class.serializer(data=obj_data)
            logger.info(f'Validating {model_name}: {obj_data}')
            if serializer.is_valid():
                try:
                    serializer.save()
                    status = 'created'
                    messages.append(f'{status} - {model_name}: {obj_data}')
                except Exception as e:
                    message = f'Error saving {model_name}: {str(e)}'
                    logger.error(message)
                    messages.append(message)
            else:
                error_str = '; '.join(
                    f'{field}: {", ".join(errors)}'
                    for field, errors in serializer.errors.items()
                )
                messages.append(
                    f'error - {model_name}: {obj_data} - {error_str}'
                )
        return {'success': True, 'messages': messages}

    def _process_excel_file(
        self,
        file,
    ) -> dict:
        """
        Process Excel file with model data.
        Checks for correct model mapping and required fields.
        Then processes data using serializers.
        Turn execel data into list of dicts and call _process_model_data.
        """
        filename = os.path.splitext(file.name)[0].lower()
        mapping_class: BaseModelMapping = self.model_mapping_class.get(
            filename
        )
        if not mapping_class:
            if filename in self.private_models_mapping_class:
                return {
                    'success': False,
                    'messages': [
                        UploadServiceMessages.RESTRICTED_UPLOAD + filename
                    ],
                }
            return {
                'success': False,
                'messages': [
                    UploadServiceMessages.UNKNOWN_MODEL_MAPPING + filename
                ],
            }
        serializer_class = mapping_class.serializer
        if serializer_class is None:
            return {
                'success': False,
                'messages': [
                    UploadServiceMessages.NO_MODEL_SERIALIZER + filename,
                ],
            }
        wb = openpyxl.load_workbook(file)
        ws = wb.active
        headers = [
            cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))
        ]
        excel_fields = set(headers)
        expected_fields = set(mapping_class.expected_fields)
        if excel_fields != expected_fields:
            if expected_fields.issuperset(excel_fields):
                missing = str(expected_fields - excel_fields)
                m = UploadServiceMessages.EXCEL_MISSING_MODEL_FIELDS + missing
            else:
                invalid_field = str(excel_fields - expected_fields)
                m = (
                    UploadServiceMessages.EXCEL_INVALID_MODEL_FIELDS
                    + invalid_field
                )
            return {
                'success': False,
                'messages': [m],
            }
        model_data = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            obj_data = dict(zip(headers, row, strict=False))
            model_data.append(mapping_class.agregate_model_fields(obj_data))
        return self._process_model_data(filename, model_data)

    def summarize_results(self, result: dict) -> dict:
        """Summarize results by counting created, and error messages."""
        messages = result.get('messages', [])
        summary = {'created': 0, 'errors': 0}
        for msg in messages:
            msg_lower = msg.lower()
            if 'created' in msg_lower:
                summary['created'] += 1
            elif 'error' in msg_lower:
                summary['errors'] += 1
        summary_messages = [
            f'Created: {summary["created"]} db objects',
            f'Errors(skipped): {summary["errors"]}',
        ]
        return {'success': summary['created'] > 0, 'message': summary_messages}
